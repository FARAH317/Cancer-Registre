"""
Générateur d'exports Excel (XLSX) avec openpyxl.
Formats : liste patients, diagnostics, traitements, rapport épidémiologique.
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# ── Couleurs thème sombre → Excel clair ──────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1A3A5C")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
ALT_FILL  = PatternFill("solid", fgColor="EBF4FF")
TITLE_FONT= Font(color="1A3A5C", bold=True, size=14, name="Calibri")
SECTION_FONT = Font(color="0077CC", bold=True, size=11, name="Calibri")
SECTION_FILL = PatternFill("solid", fgColor="E8F4FF")
THIN_BORDER = Border(
    left=Side(style='thin', color='CCDDEE'),
    right=Side(style='thin', color='CCDDEE'),
    top=Side(style='thin', color='CCDDEE'),
    bottom=Side(style='thin', color='CCDDEE'),
)

def _style_header_row(ws, row, cols):
    """Applique le style header à une ligne."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def _style_data_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = ALT_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')

def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


# ─────────────────────────────────────────────────────────────────
# 1. EXPORT LISTE PATIENTS
# ─────────────────────────────────────────────────────────────────

def export_patients_xlsx(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"
    ws.sheet_view.showGridLines = True

    # Titre
    ws.merge_cells('A1:N1')
    ws['A1'] = f"Liste des Patients — RegistreCancer.dz — {date.today().strftime('%d/%m/%Y')}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # En-têtes
    headers = [
        'N° Registre', 'Nom', 'Prénom', 'Date Naiss.', 'Âge Diag.',
        'Sexe', 'Wilaya', 'Commune', 'Téléphone',
        'Statut dossier', 'Statut vital', 'Date enreg.',
        'Médecin réf.', 'Établissement',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    # Données
    STATUT_LABELS = {
        'nouveau':'Nouveau', 'traitement':'En traitement', 'remission':'Rémission',
        'perdu':'Perdu de vue', 'decede':'Décédé', 'archive':'Archivé',
    }
    VITAL_LABELS = {'vivant':'Vivant', 'decede':'Décédé', 'inconnu':'Inconnu'}

    for i, p in enumerate(queryset):
        row = i + 3
        ws.cell(row=row, column=1,  value=p.registration_number)
        ws.cell(row=row, column=2,  value=p.nom)
        ws.cell(row=row, column=3,  value=p.prenom)
        ws.cell(row=row, column=4,  value=p.date_naissance.strftime('%d/%m/%Y') if p.date_naissance else '')
        ws.cell(row=row, column=5,  value=p.age_diagnostic)
        ws.cell(row=row, column=6,  value={'M':'Masculin','F':'Féminin','U':'Inconnu'}.get(p.sexe, p.sexe))
        ws.cell(row=row, column=7,  value=p.wilaya)
        ws.cell(row=row, column=8,  value=p.commune)
        ws.cell(row=row, column=9,  value=p.telephone or '')
        ws.cell(row=row, column=10, value=STATUT_LABELS.get(p.statut_dossier, p.statut_dossier))
        ws.cell(row=row, column=11, value=VITAL_LABELS.get(p.statut_vital, p.statut_vital))
        ws.cell(row=row, column=12, value=p.date_enregistrement.strftime('%d/%m/%Y') if p.date_enregistrement else '')
        ws.cell(row=row, column=13, value=str(p.medecin_referent) if p.medecin_referent else '')
        ws.cell(row=row, column=14, value=p.etablissement_prise_en_charge or '')
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
# 2. EXPORT DIAGNOSTICS
# ─────────────────────────────────────────────────────────────────

def export_diagnostics_xlsx(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostics"

    ws.merge_cells('A1:L1')
    ws['A1'] = f"Liste des Diagnostics — RegistreCancer.dz — {date.today().strftime('%d/%m/%Y')}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = [
        'N° Registre', 'Patient', 'Date Diag.',
        'Topographie (code)', 'Topographie (libellé)',
        'Morphologie (code)', 'Morphologie (libellé)',
        'Stade AJCC', 'TNM – T', 'TNM – N', 'TNM – M',
        'Base diagnostic',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    for i, d in enumerate(queryset):
        row = i + 3
        ws.cell(row=row, column=1,  value=d.patient.registration_number)
        ws.cell(row=row, column=2,  value=d.patient.get_full_name())
        ws.cell(row=row, column=3,  value=d.date_diagnostic.strftime('%d/%m/%Y') if d.date_diagnostic else '')
        ws.cell(row=row, column=4,  value=d.topographie_code)
        ws.cell(row=row, column=5,  value=d.topographie_libelle)
        ws.cell(row=row, column=6,  value=d.morphologie_code)
        ws.cell(row=row, column=7,  value=d.morphologie_libelle)
        ws.cell(row=row, column=8,  value=d.stade_ajcc)
        ws.cell(row=row, column=9,  value=d.tnm_t)
        ws.cell(row=row, column=10, value=d.tnm_n)
        ws.cell(row=row, column=11, value=d.tnm_m)
        ws.cell(row=row, column=12, value=d.base_diagnostic)
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
# 3. RAPPORT ÉPIDÉMIOLOGIQUE (multi-feuilles avec stats)
# ─────────────────────────────────────────────────────────────────

def export_rapport_epidemio_xlsx(stats_data, annee=None):
    wb = Workbook()

    # ── Feuille de couverture ──────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Rapport"
    ws_cover.sheet_view.showGridLines = False

    ws_cover.merge_cells('A1:F1')
    ws_cover['A1'] = "REGISTRE DU CANCER — ALGÉRIE"
    ws_cover['A1'].font = Font(color="1A3A5C", bold=True, size=18, name="Calibri")
    ws_cover['A1'].alignment = Alignment(horizontal='center')

    ws_cover.merge_cells('A2:F2')
    ws_cover['A2'] = f"Rapport épidémiologique{' — ' + str(annee) if annee else ''}"
    ws_cover['A2'].font = Font(color="0077CC", bold=True, size=13, name="Calibri")
    ws_cover['A2'].alignment = Alignment(horizontal='center')

    ws_cover.merge_cells('A3:F3')
    ws_cover['A3'] = f"Généré le {date.today().strftime('%d %B %Y')}"
    ws_cover['A3'].font = Font(color="666666", size=10, name="Calibri")
    ws_cover['A3'].alignment = Alignment(horizontal='center')

    # KPIs
    row = 5
    kpis = stats_data.get('kpis', {})
    kpi_data = [
        ('Total patients',      kpis.get('total_patients', 0)),
        ('Total diagnostics',   kpis.get('total_diagnostics', 0)),
        ('Total traitements',   kpis.get('total_traitements', 0)),
        ('En traitement',       kpis.get('en_traitement', 0)),
        ('En rémission',        kpis.get('en_remission', 0)),
        ('Décédés',             kpis.get('decedes', 0)),
    ]
    ws_cover.merge_cells(f'A{row}:F{row}')
    ws_cover[f'A{row}'] = "INDICATEURS CLÉS"
    ws_cover[f'A{row}'].font = SECTION_FONT
    ws_cover[f'A{row}'].fill = SECTION_FILL
    ws_cover[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1

    for label, val in kpi_data:
        ws_cover.cell(row=row, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws_cover.cell(row=row, column=2, value=val).font = Font(color="0077CC", bold=True, size=12, name="Calibri")
        _style_data_row(ws_cover, row, 6, alt=(row % 2 == 0))
        row += 1

    _auto_width(ws_cover)

    # ── Feuille Incidence par wilaya ────────────────────────────
    ws_wil = wb.create_sheet("Incidence par wilaya")
    ws_wil.merge_cells('A1:C1')
    ws_wil['A1'] = "Incidence par Wilaya"
    ws_wil['A1'].font = TITLE_FONT
    ws_wil['A1'].alignment = Alignment(horizontal='center')
    ws_wil.row_dimensions[1].height = 24

    for col, h in enumerate(['Wilaya', 'Nombre de patients', 'Rang'], 1):
        ws_wil.cell(row=2, column=col, value=h)
    _style_header_row(ws_wil, 2, 3)

    wilayas = stats_data.get('top_wilayas', [])
    for i, w in enumerate(wilayas):
        row = i + 3
        ws_wil.cell(row=row, column=1, value=w.get('wilaya', ''))
        ws_wil.cell(row=row, column=2, value=w.get('count', 0))
        ws_wil.cell(row=row, column=3, value=i + 1)
        _style_data_row(ws_wil, row, 3, alt=(i % 2 == 1))

    # Graphique bar
    if wilayas:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Patients par wilaya"
        chart.y_axis.title = "Wilaya"
        chart.x_axis.title = "Nombre"
        chart.style = 10
        chart.width = 20
        chart.height = 14
        data_ref  = Reference(ws_wil, min_col=2, min_row=2, max_row=min(len(wilayas)+2, 17))
        cats_ref  = Reference(ws_wil, min_col=1, min_row=3, max_row=min(len(wilayas)+2, 17))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_wil.add_chart(chart, "E3")

    _auto_width(ws_wil)

    # ── Feuille Top cancers ──────────────────────────────────────
    ws_can = wb.create_sheet("Top cancers")
    ws_can.merge_cells('A1:D1')
    ws_can['A1'] = "Top Localisations Tumorales (ICD-O-3)"
    ws_can['A1'].font = TITLE_FONT
    ws_can['A1'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(['Code', 'Localisation', 'Nombre de cas', 'Rang'], 1):
        ws_can.cell(row=2, column=col, value=h)
    _style_header_row(ws_can, 2, 4)

    cancers = stats_data.get('top_cancers', [])
    for i, c in enumerate(cancers):
        row = i + 3
        ws_can.cell(row=row, column=1, value=c.get('topographie_code', ''))
        ws_can.cell(row=row, column=2, value=c.get('topographie_libelle', ''))
        ws_can.cell(row=row, column=3, value=c.get('count', 0))
        ws_can.cell(row=row, column=4, value=i + 1)
        _style_data_row(ws_can, row, 4, alt=(i % 2 == 1))

    _auto_width(ws_can)

    # ── Feuille Distribution par stade ──────────────────────────
    ws_stade = wb.create_sheet("Stades AJCC")
    ws_stade.merge_cells('A1:B1')
    ws_stade['A1'] = "Distribution par Stade AJCC/UICC"
    ws_stade['A1'].font = TITLE_FONT
    ws_stade['A1'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(['Stade', 'Nombre de cas'], 1):
        ws_stade.cell(row=2, column=col, value=h)
    _style_header_row(ws_stade, 2, 2)

    stades = [s for s in stats_data.get('par_stade', []) if s.get('count', 0) > 0]
    for i, s in enumerate(stades):
        row = i + 3
        label = 'Inconnu' if s['stade_ajcc'] == 'U' else f"Stade {s['stade_ajcc']}"
        ws_stade.cell(row=row, column=1, value=label)
        ws_stade.cell(row=row, column=2, value=s.get('count', 0))
        _style_data_row(ws_stade, row, 2, alt=(i % 2 == 1))

    if stades:
        pie = PieChart()
        pie.title = "Répartition par stade AJCC"
        pie.style  = 10
        pie.width  = 16
        pie.height = 14
        d_ref = Reference(ws_stade, min_col=2, min_row=2, max_row=len(stades)+2)
        c_ref = Reference(ws_stade, min_col=1, min_row=3, max_row=len(stades)+2)
        pie.add_data(d_ref, titles_from_data=True)
        pie.set_categories(c_ref)
        ws_stade.add_chart(pie, "D3")

    _auto_width(ws_stade)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
# 4. EXPORT CANREG5 (format CSV compatible)
# ─────────────────────────────────────────────────────────────────

def export_canreg5_csv(patient_queryset, diagnostic_queryset):
    """
    Format CSV compatible CanReg5 (IARC).
    Colonnes standardisées selon le dictionnaire CanReg5.
    """
    import csv
    import io as _io

    output = _io.StringIO()
    writer = csv.writer(output, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # En-tête CanReg5 standard
    writer.writerow([
        'REGISTRATIONNUMBER',
        'TUMOURREGISTRATIONNUMBER',
        'SEX',
        'BIRTHDATE',
        'INCIDENCEDATE',
        'TOPOGRAPHY',
        'MORPHOLOGY',
        'BEHAVIOUR',
        'GRADE',
        'BASIS',
        'LATERALITY',
        'TNMT',
        'TNMN',
        'TNMM',
        'STAGEVITAL',
        'DATEOFVITALSTATUS',
        'SOURCE',
        'COMMENTS',
    ])

    # Construire un mapping diagnostic par patient
    diag_map = {}
    for d in diagnostic_queryset:
        if d.patient_id not in diag_map:
            diag_map[d.patient_id] = []
        diag_map[d.patient_id].append(d)

    SEX_MAP   = {'M': '1', 'F': '2', 'U': '9'}
    BASIS_MAP = {
        'clinique': '1', 'imagerie': '2', 'cytologie': '5',
        'histologie': '7', 'autopsie': '8', 'dco': '0', 'autre': '9',
    }
    VITAL_MAP = {'vivant': '1', 'decede': '2', 'inconnu': '9'}

    tumour_num = 1
    for patient in patient_queryset:
        diags = diag_map.get(patient.id, [None])
        for diag in diags:
            row = [
                patient.registration_number or '',
                str(tumour_num).zfill(4),
                SEX_MAP.get(patient.sexe, '9'),
                patient.date_naissance.strftime('%Y%m%d') if patient.date_naissance else '99999999',
                diag.date_diagnostic.strftime('%Y%m%d') if diag and diag.date_diagnostic else '99999999',
                (diag.topographie_code or '').replace('C', '').replace('.', '') if diag else '999',
                (diag.morphologie_code or '').replace('/', '') if diag else '9999',
                '3',  # Behaviour: 3=malignant
                '9',  # Grade: unknown
                BASIS_MAP.get(getattr(diag, 'base_diagnostic', ''), '9') if diag else '9',
                '0',  # Laterality: NA
                (diag.tnm_t or '') if diag else '',
                (diag.tnm_n or '') if diag else '',
                (diag.tnm_m or '') if diag else '',
                VITAL_MAP.get(patient.statut_vital, '9'),
                '',   # DateOfVitalStatus
                'REG',
                '',
            ]
            writer.writerow(row)
            tumour_num += 1

    output.seek(0)
    return output.getvalue().encode('utf-8')
